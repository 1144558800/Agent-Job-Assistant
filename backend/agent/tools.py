# -*- coding: utf-8 -*-
"""
Agent 工具定义 - 将现有功能封装为 LangChain Tools
"""
import os
import json
import random
import asyncio
import threading
from typing import List, Optional
from datetime import datetime
from loguru import logger
from langchain_core.tools import tool
from langchain_core.runnables import RunnableConfig

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import UPLOAD_DIR, SCRAPED_DATA_DIR
from scrapers.scraper_manager import ScraperManager
from rag.faiss_store import FaissStore
from rag.document_processor import DocumentProcessor
from rag.embeddings import EmbeddingService
from rag.qa_engine import QAEngine
from resume.parser import extract_resume_text
from resume.matcher import ResumeMatcher
from resume.resume_editor import (
    polish_resume_with_ai,
    generate_polished_docx,
    generate_polished_pdf,
    polish_docx_inplace,
)
from agent.apply_manager import (
    generate_greeting, is_already_applied, record_apply, get_apply_statistics
)
from agent.hermes_memory import get_hermes_memory
from agent.guardrails import get_guardrail_manager


# ---- 全局单例（延迟初始化） ----
_scraper_manager: Optional[ScraperManager] = None
_faiss_store: Optional[FaissStore] = None
_doc_processor: Optional[DocumentProcessor] = None
_embedding_service: Optional[EmbeddingService] = None
_qa_engine: Optional[QAEngine] = None


def _get_scraper_manager() -> ScraperManager:
    global _scraper_manager
    if _scraper_manager is None:
        _scraper_manager = ScraperManager()
    return _scraper_manager


def _get_faiss_store() -> FaissStore:
    global _faiss_store
    if _faiss_store is None:
        from config import FAISS_INDEX_DIR
        _faiss_store = FaissStore(str(FAISS_INDEX_DIR))
        try:
            _faiss_store.load()
        except FileNotFoundError:
            pass
    return _faiss_store


def _get_doc_processor() -> DocumentProcessor:
    global _doc_processor
    if _doc_processor is None:
        _doc_processor = DocumentProcessor()
    return _doc_processor


def _get_embedding_service() -> EmbeddingService:
    global _embedding_service
    if _embedding_service is None:
        _embedding_service = EmbeddingService()
    return _embedding_service


def _get_qa_engine() -> QAEngine:
    global _qa_engine
    if _qa_engine is None:
        _qa_engine = QAEngine()
    return _qa_engine


# ---- 工具定义 ----

# ---- 自动重试机制 ----
# 只对网络/超时类瞬态错误重试，逻辑错误（如"无结果"）不重试

_RETRYABLE_KEYWORDS = [
    "timeout", "超时", "timed out",
    "connection", "连接", "connect",
    "network", "网络",
    "temporary", "临时",
    "unreachable", "不可达",
    "rate limit", "限流",
    "too many requests", "请求过多",
    "server error", "服务器错误", "500",
    "service unavailable", "503",
    "gateway", "502", "504",
    "reset by peer", "连接重置",
    "eof", "econnreset", "econnrefused",
    "proxy", "代理",
    "dns", "域名解析",
]

_MAX_RETRIES = 2  # 最多重试2次
_RETRY_BASE_DELAY = 2.0  # 基础等待秒数（指数退避：2s, 4s）


def _is_retryable_error(error_msg: str) -> bool:
    """判断错误是否为可重试的瞬态错误（网络/超时类）"""
    if not error_msg:
        return False
    msg_lower = error_msg.lower()
    for keyword in _RETRYABLE_KEYWORDS:
        if keyword in msg_lower:
            return True
    return False


def _is_logic_error(error_msg: str) -> bool:
    """判断是否为不应重试的逻辑错误"""
    non_retryable = [
        "没有搜索到", "未搜索到", "未找到", "not found",
        "没有可", "没有匹配", "不支持", "invalid",
        "格式", "文件", "简历", "校验",
    ]
    msg_lower = error_msg.lower() if error_msg else ""
    return any(kw in msg_lower for kw in non_retryable)


async def _retry_tool(tool_name: str, coro_func, *args, **kwargs):
    """对工具调用执行自动重试（仅对瞬态错误重试）
    
    参数:
        tool_name: 工具名称（用于日志）
        coro_func: 要执行的异步函数
        *args, **kwargs: 传递给 coro_func 的参数
    
    返回 coro_func 的执行结果
    """
    last_error = None
    for attempt in range(_MAX_RETRIES + 1):  # 0, 1, 2（总共最多执行3次）
        try:
            if attempt > 0:
                delay = _RETRY_BASE_DELAY * (2 ** (attempt - 1))  # 指数退避
                logger.warning("[Retry][{}] 第{}次重试，等待{:.1f}秒...", tool_name, attempt, delay)
                await asyncio.sleep(delay)
            
            result = await coro_func(*args, **kwargs)
            
            # 检查返回的 JSON 结果是否包含可重试的错误
            if isinstance(result, str):
                try:
                    parsed = json.loads(result)
                except json.JSONDecodeError:
                    # 非 JSON 字符串，直接返回
                    return result
                
                if not parsed.get("success") and attempt < _MAX_RETRIES:
                    error_msg = parsed.get("message", "")
                    if _is_retryable_error(error_msg) and not _is_logic_error(error_msg):
                        logger.warning("[Retry][{}] 检测到可重试错误: {}，准备重试(第{}/{})",
                                     tool_name, error_msg[:100], attempt + 1, _MAX_RETRIES)
                        last_error = error_msg
                        continue
            
            if attempt > 0:
                logger.info("[Retry][{}] 重试成功（第{}次尝试）", tool_name, attempt + 1)
            return result
            
        except Exception as e:
            error_msg = str(e)
            if attempt < _MAX_RETRIES and _is_retryable_error(error_msg):
                logger.warning("[Retry][{}] 异常(可重试): {}，准备重试(第{}/{})",
                             tool_name, error_msg[:100], attempt + 1, _MAX_RETRIES)
                last_error = error_msg
                continue
            else:
                # 非可重试异常或已达最大重试次数，直接抛出
                if attempt >= _MAX_RETRIES:
                    logger.error("[Retry][{}] 已达最大重试次数({})，最终错误: {}",
                               tool_name, _MAX_RETRIES, last_error or error_msg[:100])
                raise
    
    # 不应到达此处，但作为安全兜底
    raise RuntimeError(f"[Retry][{tool_name}] 重试耗尽，最终错误: {last_error}")

# ---- 简历文件查找（支持对话隔离） ----

def _find_resume_files(thread_id: Optional[str] = None) -> list:
    """
    查找简历文件，优先扫描对话专属目录（UPLOAD_DIR/{thread_id}/）。
    如果 thread_id 为 "default" 或对话目录不存在，回退到全局目录（兼容旧逻辑）。
    实现不同对话之间的简历隔离。
    
    参数:
        thread_id: 当前对话的线程ID
    
    返回: 按修改时间排序的简历文件路径列表（最新的在前）
    """
    resume_files = []
    
    # 1. 优先扫描对话专属目录
    if thread_id and thread_id != "default":
        session_dir = UPLOAD_DIR / thread_id
        if session_dir.exists():
            for ext in [".pdf", ".docx", ".doc", ".txt"]:
                resume_files.extend(list(session_dir.glob(f"*{ext}")))
            if resume_files:
                logger.info("[Tools][_find_resume_files] 在对话专属目录找到 {} 个文件 (thread={})",
                           len(resume_files), thread_id)
                resume_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
                return resume_files
            else:
                # 对话目录存在但为空，不给回退，需要用户在该对话中上传简历
                logger.info("[Tools][_find_resume_files] 对话专属目录为空 (thread={})，不回退到全局", thread_id)
                return []
    
    # 2. 仅对 "default" 或无 thread_id 的情况回退到全局目录（兼容旧逻辑）
    if not thread_id or thread_id == "default":
        if UPLOAD_DIR.exists():
            for ext in [".pdf", ".docx", ".doc", ".txt"]:
                resume_files.extend(list(UPLOAD_DIR.glob(f"*{ext}")))
            if resume_files:
                logger.info("[Tools][_find_resume_files] 在全局目录找到 {} 个文件", len(resume_files))
    
    resume_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return resume_files


@tool
async def search_jobs(keyword: str, city: str, platforms: Optional[str] = None) -> str:
    """
    在招聘平台上搜索岗位。支持多平台并行搜索。
    
    参数:
        keyword: 搜索关键词，如 "Python开发"、"数据分析"
        city: 城市名称，如 "北京"、"上海"、"深圳"
        platforms: 可选，指定平台，用逗号分隔，如 "boss,zhaopin,liepin,job51"。不指定则搜索全部平台。
    """
    logger.info("[Tools][search_jobs] === 被调用 ===")
    logger.info("[Tools][search_jobs] 入参: keyword={}, city={}, platforms={}", keyword, city, platforms)
    
    # ---- 城市名白名单校验（防止 Agent 幻觉传入非法城市名） ----
    VALID_CITIES = {
        "北京", "上海", "广州", "深圳", "杭州", "南京", "成都", "武汉", "西安",
        "长沙", "重庆", "苏州", "东莞", "合肥", "郑州", "济南", "青岛", "厦门",
        "福州", "大连", "沈阳", "哈尔滨", "昆明", "贵阳", "南宁", "海口", "石家庄",
        "太原", "呼和浩特", "兰州", "银川", "西宁", "乌鲁木齐", "拉萨", "南昌",
        "珠海", "佛山", "中山", "惠州", "宁波", "温州", "无锡", "常州", "南通",
        "嘉兴", "绍兴", "金华", "台州", "泉州", "漳州", "徐州", "扬州", "泰州",
        "镇江", "盐城", "淮安", "连云港", "宿迁", "烟台", "威海", "潍坊", "淄博",
        "临沂", "济宁", "泰安", "洛阳", "新乡", "南阳", "许昌", "宜昌", "襄阳",
        "株洲", "湘潭", "衡阳", "岳阳", "绵阳", "德阳", "宜宾", "柳州", "桂林",
        "三亚", "海口", "拉萨", "银川", "西宁", "呼和浩特", "乌鲁木齐",
        # 支持带"市"后缀
        "北京市", "上海市", "广州市", "深圳市", "杭州市", "南京市", "成都市",
        "武汉市", "西安市", "长沙市", "重庆市", "苏州市", "东莞市", "合肥市",
    }
    # 自动去除"市"后缀进行匹配
    city_normalized = city.rstrip("市") if city.endswith("市") and city[:-1] in VALID_CITIES else city
    if city not in VALID_CITIES and city_normalized not in VALID_CITIES:
        # 尝试模糊匹配：检查是否以某个合法城市名开头
        matched = None
        for vc in VALID_CITIES:
            if len(vc) >= 2 and city.startswith(vc):
                matched = vc
                break
        if matched:
            city = matched
            logger.warning("[Tools][search_jobs] 城市名自动修正: '{}' -> '{}'", city_normalized, matched)
        else:
            logger.warning("[Tools][search_jobs] 城市名不在白名单中: '{}'", city)
            return json.dumps({
                "success": False,
                "message": f"城市名 '{city}' 不合法，请使用标准中国城市名称，如：北京、上海、广州、深圳、杭州、成都、武汉等"
            }, ensure_ascii=False)
    
    # ---- 关键词校验 ----
    if not keyword or not keyword.strip():
        return json.dumps({
            "success": False,
            "message": "搜索关键词不能为空，请提供岗位名称"
        }, ensure_ascii=False)
    
    mgr = _get_scraper_manager()
    
    # 平台名称映射（Agent可能传入的别名 -> 爬虫内部key）
    PLATFORM_ALIAS_MAP = {
        "job51": "51job",
        "51job": "51job",
        "前程无忧": "51job",
        "boss": "boss",
        "boss直聘": "boss",
        "zhaopin": "zhaopin",
        "智联": "zhaopin",
        "智联招聘": "zhaopin",
        "liepin": "liepin",
        "猎聘": "liepin",
    }
    
    platform_list = None
    if platforms:
        raw_list = [p.strip() for p in platforms.split(",")]
        platform_list = [PLATFORM_ALIAS_MAP.get(p, p) for p in raw_list]
        logger.info("[Tools][search_jobs] 原始平台: {} -> 映射后: {}", raw_list, platform_list)
    else:
        logger.info("[Tools][search_jobs] 未指定平台，将搜索全部4个平台")
    
    try:
        # 自动重试：将搜索调用包装在重试机制中（网络/超时类错误自动重试最多2次）
        async def _do_search():
            return await mgr.search_all(keyword, city, platforms=platform_list)
        
        jobs = await _retry_tool("search_jobs", _do_search)
        logger.info("[Tools][search_jobs] search_all 返回: {} 个岗位", len(jobs))
        
        if not jobs:
            logger.warning("[Tools][search_jobs] 搜索结果为0！请检查日志中的[ScraperManager]和各平台日志。")
            return json.dumps(
                {"success": True, "message": f"未搜索到 {city} 的 {keyword} 岗位，请查看日志排查原因。", "count": 0, "jobs": []},
                ensure_ascii=False
            )
        
        # 只返回摘要信息，完整数据保留在内存中
        sample_jobs = [j.model_dump() for j in jobs[:5]]
        summary = {
            "success": True,
            "message": f"成功搜索到 {len(jobs)} 个岗位",
            "count": len(jobs),
            "summary_by_platform": {},
            "sample_jobs": sample_jobs
        }
        
        for j in jobs:
            p = j.platform or "未知"
            summary["summary_by_platform"][p] = summary["summary_by_platform"].get(p, 0) + 1
        
        # Hermes 自进化：记录搜索经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="search",
                success=len(jobs) > 0,
                result_summary=f"搜索到 {len(jobs)} 个岗位" if jobs else "未搜索到岗位",
                input_params={
                    "keyword": keyword,
                    "city": city,
                    "platforms": platforms or "全部",
                },
                result_detail={
                    "count": len(jobs),
                    "platform_distribution": summary["summary_by_platform"],
                },
            )
            # 检查自动反思
            from config import HERMES_AUTO_REFLECT_THRESHOLD
            if HERMES_AUTO_REFLECT_THRESHOLD > 0:
                stats = hm.get_statistics()
                search_count = stats.get("by_action", {}).get("search", {}).get("total", 0)
                if search_count > 0 and search_count % HERMES_AUTO_REFLECT_THRESHOLD == 0:
                    logger.info("[Tools] Hermes 自动反思触发: 已累积 {} 条搜索经验", search_count)
                    hm.reflect()
        except Exception as e:
            logger.warning("[Tools] Hermes 经验记录失败: {}", e)
        
        logger.info("[Tools][search_jobs] 返回摘要: 总数={}, 平台分布={}", len(jobs), summary["summary_by_platform"])
        return json.dumps(summary, ensure_ascii=False)
    except Exception as e:
        logger.error("[Tools][search_jobs] 异常: type={}, msg={}", type(e).__name__, str(e))
        import traceback
        logger.error("[Tools][search_jobs] 异常堆栈:\n{}", traceback.format_exc())
        return json.dumps({"success": False, "message": f"搜索失败: {str(e)}", "count": 0}, ensure_ascii=False)


@tool
async def save_to_knowledge(job_ids: Optional[str] = None) -> str:
    """
    将搜索到的岗位保存到 FAISS 向量知识库，以便后续 RAG 查询和分析。
    如果不指定 job_ids，则保存所有搜索到的岗位。
    
    参数:
        job_ids: 可选，要保存的岗位ID列表（用逗号分隔）。不指定则保存全部。
    """
    logger.info(f"[Agent Tool] save_to_knowledge: job_ids={job_ids}")
    
    mgr = _get_scraper_manager()
    store = _get_faiss_store()
    doc_processor = _get_doc_processor()
    embed_service = _get_embedding_service()
    
    try:
        # 从 scraper_manager 获取最近的搜索结果
        # 注意：需要从 Agent 状态中获取，这里暂时通过内部属性获取
        all_jobs = getattr(mgr, '_last_search_results', [])
        
        if not all_jobs:
            return json.dumps({"success": False, "message": "没有可保存的岗位数据，请先搜索岗位。", "count": 0}, ensure_ascii=False)
        
        # 筛选指定岗位
        if job_ids:
            id_list = set(job_ids.split(","))
            jobs = [j for j in all_jobs if getattr(j, 'job_url', '') and any(jid in getattr(j, 'job_url', '') for jid in id_list)]
        else:
            jobs = all_jobs
        
        if not jobs:
            return json.dumps({"success": False, "message": "没有匹配的岗位数据。", "count": 0}, ensure_ascii=False)
        
        # 文档处理和向量化
        documents = doc_processor.jobs_to_documents(jobs)
        chunks = doc_processor.split_documents(documents)
        texts = [chunk["content"] for chunk in chunks]
        vectors = embed_service.embed_texts(texts)
        
        # 确保索引已加载或创建
        if store.index is None:
            try:
                store.load()
            except FileNotFoundError:
                pass
            if store.index is None:
                dim = embed_service.get_embedding_dimension()
                store.create_index(dim)
        
        # 添加到 FAISS
        if vectors:
            store.add_vectors(vectors, chunks)
            store.save()
        
        return json.dumps({
            "success": True,
            "message": f"成功保存 {len(jobs)} 个岗位（{len(chunks)} 个文档块）到知识库",
            "count": len(jobs),
            "chunk_count": len(chunks),
            "faiss_total": store.total_count
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] save_to_knowledge 失败: {e}")
        return json.dumps({"success": False, "message": f"保存失败: {str(e)}"}, ensure_ascii=False)


@tool
async def query_knowledge(question: str) -> str:
    """
    基于已保存的岗位知识库进行 RAG 问答。可以查询岗位要求、公司信息、薪资水平等。
    
    参数:
        question: 要查询的问题，如 "有哪些公司提供远程办公"、"Python岗位的平均薪资"
    """
    logger.info(f"[Agent Tool] query_knowledge: {question}")
    
    store = _get_faiss_store()
    
    try:
        # 确保索引已加载
        if store.index is None:
            try:
                store.load()
            except FileNotFoundError:
                pass
        
        if store.index is None or store.total_count == 0:
            return json.dumps({"success": False, "message": "知识库为空，请先搜索并保存岗位数据。", "answer": ""}, ensure_ascii=False)
        
        qa_engine = _get_qa_engine()
        embed_service = _get_embedding_service()
        
        # 生成查询向量
        query_vector = embed_service.embed_text(question)
        
        # FAISS 检索
        search_results = store.search(query_vector, top_k=5)
        
        # RAG 问答
        answer = qa_engine.get_answer(question, search_results)
        
        # Hermes 自进化：记录知识库查询经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="query",
                success=True,
                result_summary=f"知识库查询: {question[:80]}",
                input_params={"question": question[:200]},
                result_detail={
                    "answer_length": len(answer) if answer else 0,
                    "reference_count": len(search_results),
                    "top_score": round(float(search_results[0].get("score", 0)), 4) if search_results else 0,
                },
            )
            logger.debug("[Tools][query_knowledge] Hermes 经验已记录")
        except Exception as e:
            logger.warning("[Tools][query_knowledge] Hermes 经验记录失败: {}", e)
        
        # 构造参考来源
        references = []
        for r in search_results[:3]:
            doc = r.get("document", {})
            meta = doc.get("metadata", {})
            references.append({
                "job_name": meta.get("job_name", "未知"),
                "company_name": meta.get("company_name", "未知"),
                "score": round(float(r.get("score", 0)), 4)
            })
        
        return json.dumps({
            "success": True,
            "answer": answer,
            "question": question,
            "references": references,
            "total_docs": store.total_count
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] query_knowledge 失败: {e}")
        # Hermes 记录失败经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="query",
                success=False,
                result_summary=f"知识库查询失败: {str(e)[:100]}",
                input_params={"question": question[:200]},
            )
        except Exception as he:
            logger.warning("[Tools][query_knowledge] Hermes 异常记录失败: {}", he)
        return json.dumps({"success": False, "message": f"查询失败: {str(e)}", "answer": ""}, ensure_ascii=False)


@tool
async def analyze_jobs(analysis_type: str = "comprehensive") -> str:
    """
    分析搜索到的岗位数据。支持薪资分析、城市分布、技能要求等维度。
    
    参数:
        analysis_type: 分析类型，可选: "salary"（薪资分析）, "city"（城市分布）, "skill"（技能要求）, "comprehensive"（综合分析）
    """
    logger.info(f"[Agent Tool] analyze_jobs: type={analysis_type}")
    
    mgr = _get_scraper_manager()
    all_jobs = getattr(mgr, '_last_search_results', [])
    
    if not all_jobs:
        return json.dumps({"success": False, "message": "没有可分析的岗位数据，请先搜索岗位。"}, ensure_ascii=False)
    
    try:
        logger.info("[Tools][analyze_jobs] 开始分析，岗位总数={}, 分析类型={}", len(all_jobs), analysis_type)
        
        result = {"success": True, "total_jobs": len(all_jobs), "analysis_type": analysis_type}
        
        # 平台分布
        platform_count = {}
        for j in all_jobs:
            p = j.platform or "未知"
            platform_count[p] = platform_count.get(p, 0) + 1
        result["platform_distribution"] = platform_count
        
        # 城市分布
        city_count = {}
        for j in all_jobs:
            c = j.location or "未知"
            city_count[c] = city_count.get(c, 0) + 1
        result["city_distribution"] = dict(sorted(city_count.items(), key=lambda x: x[1], reverse=True)[:10])
        
        # 薪资分析
        salaries = []
        salary_count = {"面议": 0, "10K以下": 0, "10-20K": 0, "20-30K": 0, "30-50K": 0, "50K以上": 0}
        for j in all_jobs:
            salary = j.salary or ""
            if not salary or "面议" in str(salary):
                salary_count["面议"] += 1
                continue
            try:
                # 尝试解析薪资范围，取中位数
                s = str(salary).lower().replace("k", "").replace("以上", "-999")
                parts = s.split("-")
                if len(parts) == 2:
                    avg = (float(parts[0]) + float(parts[1])) / 2
                    salaries.append(avg)
            except:
                salary_count["面议"] += 1
        
        if salaries:
            salaries.sort()
            salary_count["10K以下"] = sum(1 for s in salaries if s < 10)
            salary_count["10-20K"] = sum(1 for s in salaries if 10 <= s < 20)
            salary_count["20-30K"] = sum(1 for s in salaries if 20 <= s < 30)
            salary_count["30-50K"] = sum(1 for s in salaries if 30 <= s < 50)
            salary_count["50K以上"] = sum(1 for s in salaries if s >= 50)
            result["salary_avg"] = round(sum(salaries) / len(salaries), 1)
            result["salary_min"] = round(min(salaries), 1) if salaries else 0
            result["salary_max"] = round(max(salaries), 1) if salaries else 0
        
        result["salary_distribution"] = salary_count
        
        # 公司Top10
        company_count = {}
        for j in all_jobs:
            comp = j.company_name or "未知"
            if comp and comp != "未知":
                company_count[comp] = company_count.get(comp, 0) + 1
        result["top_companies"] = dict(sorted(company_count.items(), key=lambda x: x[1], reverse=True)[:10])
        
        # Hermes 自进化：记录分析经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="analyze",
                success=True,
                result_summary=f"分析了 {len(all_jobs)} 个岗位的{analysis_type}维度，平均薪资约{result.get('salary_avg', 'N/A')}K",
                input_params={"analysis_type": analysis_type, "total_jobs": len(all_jobs)},
                result_detail={
                    "platform_distribution": platform_count,
                    "salary_distribution": result.get("salary_distribution", {}),
                    "top_platform": max(platform_count, key=platform_count.get) if platform_count else "未知",
                },
            )
            logger.debug("[Tools][analyze_jobs] Hermes 经验已记录")
        except Exception as e:
            logger.warning("[Tools][analyze_jobs] Hermes 经验记录失败: {}", e)
        
        logger.info("[Tools][analyze_jobs] 分析完成，返回结果")
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] analyze_jobs 失败: {e}")
        # Hermes 记录失败经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="analyze",
                success=False,
                result_summary=f"分析失败: {str(e)[:100]}",
                input_params={"analysis_type": analysis_type},
            )
        except Exception as he:
            logger.warning("[Tools][analyze_jobs] Hermes 异常记录失败: {}", he)
        return json.dumps({"success": False, "message": f"分析失败: {str(e)}"}, ensure_ascii=False)


@tool
async def match_resume(config: RunnableConfig) -> str:
    """
    将已上传的简历与知识库中的岗位进行匹配，返回匹配度最高的岗位列表。
    使用前请先确保已上传简历文件。
    """
    thread_id = config.get("configurable", {}).get("thread_id", "default") if config else "default"
    logger.info("[Agent Tool] match_resume (thread={})", thread_id)
    
    mgr = _get_scraper_manager()
    all_jobs = getattr(mgr, '_last_search_results', [])
    
    if not all_jobs:
        return json.dumps({"success": False, "message": "没有可匹配的岗位数据，请先搜索岗位或从知识库加载。"}, ensure_ascii=False)
    
    # 查找已上传的简历文件（优先对话专属目录，确保对话隔离）
    resume_files = _find_resume_files(thread_id)
    
    if not resume_files:
        return json.dumps({
            "success": False,
            "message": "未找到上传的简历文件。请先通过文件上传功能上传简历（支持 PDF/Word/TXT 格式）。"
        }, ensure_ascii=False)
    
    try:
        resume_file = resume_files[-1]  # 使用最新的简历
        logger.info("[Tools][match_resume] 开始匹配，简历文件={}, 岗位总数={}", resume_file.name, len(all_jobs))
        
        # 解析简历文本
        resume_text = extract_resume_text(str(resume_file))
        if not resume_text:
            return json.dumps({
                "success": False,
                "message": "简历解析失败，请检查文件格式（支持 PDF/Word/TXT）"
            }, ensure_ascii=False)
        
        matcher = ResumeMatcher()
        
        # 使用匹配器匹配简历与岗位（转换为 dict 列表），带自动重试（API调用可能超时）
        job_dicts = [j.model_dump() for j in all_jobs]
        
        async def _do_match():
            return matcher.match_resume_with_jobs(resume_text, job_dicts)
        
        match_result = await _retry_tool("match_resume", _do_match)
        
        # Hermes 自进化：记录匹配经验
        match_success = isinstance(match_result, str) or match_result is not None
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="match",
                success=match_success,
                result_summary=f"简历与 {len(all_jobs)} 个岗位匹配完成" if match_success else "匹配未产生有效结果",
                input_params={
                    "resume_file": resume_file.name,
                    "total_jobs": len(all_jobs),
                },
                result_detail={
                    "resume_length": len(resume_text),
                    "match_type": "ai" if isinstance(match_result, str) else "keyword",
                },
            )
            logger.debug("[Tools][match_resume] Hermes 经验已记录")
        except Exception as e:
            logger.warning("[Tools][match_resume] Hermes 经验记录失败: {}", e)
        
        # 如果匹配器返回了字符串，直接返回
        if isinstance(match_result, str):
            return json.dumps({
                "success": True,
                "message": f"简历与 {len(all_jobs)} 个岗位匹配完成",
                "resume_file": resume_file.name,
                "match_result": match_result[:2000],
                "total_jobs": len(all_jobs)
            }, ensure_ascii=False)
        
        return json.dumps({
            "success": True,
            "message": f"简历与 {len(all_jobs)} 个岗位匹配完成",
            "resume_file": resume_file.name,
            "total_jobs": len(all_jobs)
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] match_resume 失败: {e}")
        # Hermes 记录失败经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="match",
                success=False,
                result_summary=f"匹配失败: {str(e)[:100]}",
                input_params={"total_jobs": len(all_jobs)},
            )
        except Exception as he:
            logger.warning("[Tools][match_resume] Hermes 异常记录失败: {}", he)
        return json.dumps({"success": False, "message": f"匹配失败: {str(e)}"}, ensure_ascii=False)


@tool
async def read_resume(config: RunnableConfig) -> str:
    """
    深度读取并解析已上传的简历文件。使用 AI 完整提取简历中的所有信息，包括基本信息、技能、
    工作经历、项目经验、教育背景、证书、求职意向等。这是进行岗位匹配前的必要步骤，
    确保系统能获取求职者的完整画像。调用后返回结构化的简历解析结果。
    """
    thread_id = config.get("configurable", {}).get("thread_id", "default") if config else "default"
    logger.info("[Agent Tool] read_resume (thread={})", thread_id)

    # 查找已上传的简历文件（优先对话专属目录，确保对话隔离）
    resume_files = _find_resume_files(thread_id)

    if not resume_files:
        logger.warning("[Tools][read_resume] 未找到简历文件")
        return json.dumps({
            "success": False,
            "message": "未找到上传的简历文件。请先上传简历（支持 PDF/Word/TXT 格式）。"
        }, ensure_ascii=False)

    try:
        resume_file = resume_files[-1]  # 使用最新的简历
        logger.info("[Tools][read_resume] 开始解析简历: {}", resume_file.name)

        # 提取文本
        resume_text = extract_resume_text(str(resume_file))
        if not resume_text:
            logger.warning("[Tools][read_resume] 简历文本提取失败")
            return json.dumps({
                "success": False,
                "message": f"简历文件 {resume_file.name} 解析失败，请检查文件格式或内容。"
            }, ensure_ascii=False)

        logger.info("[Tools][read_resume] 简历文本提取成功, 长度={} 字符", len(resume_text))

        # 使用 DeepSeek AI 深度解析简历
        qa = _get_qa_engine()
        qa._init_client()

        system_prompt = """你是一位专业的简历解析专家和职业规划师。请仔细阅读以下简历全文，逐句分析，
深入挖掘简历中的每一处信息，不要遗漏任何细节。

请按以下结构输出完整分析结果（使用 Markdown 格式，务必详实，不要概括过度）：

## 一、基本信息
- 姓名、性别、年龄、所在城市
- 最高学历、毕业院校、专业
- 工作年限（精确到年）
- 当前/最近职位
- 期望城市、期望薪资（如有）

## 二、专业技能矩阵
请用表格列出所有技能，分类为：编程语言、框架工具、数据库、云原生、AI/ML、其他
每项标注：技能名称、熟练程度（精通/熟练/了解）、来源（从简历何处提取）

## 三、工作经历（逐条详述）
对每一段工作经历，请提取：
- 公司名称、职位、起止时间
- 主要职责（逐条列出，保留原文关键描述）
- 技术栈和工具
- 取得的成果/业绩（如有数据请保留数据）

## 四、项目经历（逐条详述）
对每一个项目，请提取：
- 项目名称、角色（独立/核心/参与）
- 项目描述（保留原文完整描述）
- 使用的技术栈（精确到具体框架/库）
- 个人贡献和成果
- 解决的问题和难点

## 五、教育背景
- 学校、专业、学历、时间
- 相关课程（如有）
- 荣誉/奖项（如有）

## 六、证书与资质
- 所有证书、认证、培训
- 语言能力（英语水平等）

## 七、求职意向归纳
根据简历内容推测：
- 目标岗位方向（如：AI Agent 开发、Python 后端、全栈开发等）
- 目标行业偏好
- 核心竞争力总结（3-5个关键词）

## 八、简历完整性评估
- 哪些信息完整、哪些信息缺失
- 简历中可能隐含但未明确写出的信息（从上下文合理推断）
- 适合投递的岗位类型建议

重要：必须逐字逐句分析原文，把简历中所有信息都提取出来。不要遗漏任何细节。"""

        async def _do_read():
            response = qa.client.chat.completions.create(
                model=qa.model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"请深度解析以下简历全文，逐句分析，提取所有信息：\n\n【简历全文】\n{resume_text}"}
                ],
                temperature=0.2,
                max_tokens=6000,
            )
            return response.choices[0].message.content

        # API 调用带自动重试
        analysis_result = await _retry_tool("read_resume", _do_read)

        logger.info("[Tools][read_resume] AI 解析完成, 结果长度={} 字符", len(analysis_result) if analysis_result else 0)

        # Hermes 记录经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="read_resume",
                success=True,
                result_summary=f"简历 {resume_file.name} 深度解析完成，{len(resume_text)}字符",
                input_params={
                    "resume_file": resume_file.name,
                    "resume_length": len(resume_text),
                },
                result_detail={
                    "analysis_length": len(analysis_result) if analysis_result else 0,
                    "sections_extracted": 8,
                },
            )
            logger.debug("[Tools][read_resume] Hermes 经验已记录")
        except Exception as he:
            logger.warning("[Tools][read_resume] Hermes 经验记录失败: {}", he)

        return json.dumps({
            "success": True,
            "message": f"简历深度解析完成",
            "resume_file": resume_file.name,
            "resume_length": len(resume_text),
            "analysis": analysis_result,
        }, ensure_ascii=False)

    except Exception as e:
        logger.error("[Tools][read_resume] 解析失败: {}", e)
        # Hermes 记录失败经验
        try:
            hm = get_hermes_memory()
            hm.record_experience(
                action_type="read_resume",
                success=False,
                result_summary=f"简历解析失败: {str(e)[:100]}",
                input_params={"resume_file": resume_file.name if resume_file else "unknown"},
            )
        except Exception as he:
            logger.warning("[Tools][read_resume] Hermes 异常记录失败: {}", he)
        return json.dumps({"success": False, "message": f"简历解析失败: {str(e)}"}, ensure_ascii=False)


@tool
async def check_login_status() -> str:
    """
    检查各招聘平台的登录/Cookie 状态。对有Cookie的平台进行实际验证（HTTP请求），确保状态准确。
    """
    logger.info("[Agent Tool] check_login_status")
    
    mgr = _get_scraper_manager()
    result = {"success": True, "platforms": {}}
    
    for platform_name, scraper in mgr.scrapers.items():
        has_cookies = hasattr(scraper, "has_cookies") and scraper.has_cookies()
        cookie_count = len(scraper.get_cookies()) if has_cookies else 0
        
        status = "未登录"
        if cookie_count > 0:
            # 实际验证Cookie有效性（四平台统一逻辑），带自动重试（HTTP验证可能超时）
            if hasattr(scraper, "verify_cookies"):
                async def _do_verify():
                    return scraper.verify_cookies()
                is_valid = await _retry_tool(f"verify_{platform_name}", _do_verify)
                status = "已登录(已验证)" if is_valid else "Cookie已过期"
            else:
                status = "已登录"
        
        # 获取平台显示名
        display_name = getattr(scraper, "platform", platform_name)
        
        status_info = {
            "platform": display_name,
            "has_cookies": has_cookies,
            "cookie_count": cookie_count,
            "status": status
        }
        result["platforms"][platform_name] = status_info
    
    return json.dumps(result, ensure_ascii=False)


@tool
async def export_excel(file_name: Optional[str] = None) -> str:
    """
    将搜索到的岗位数据导出为 Excel 文件。文件保存在 uploads 目录。
    
    参数:
        file_name: 可选，导出文件名（不含扩展名），默认为 "岗位数据_日期时间"
    """
    logger.info(f"[Agent Tool] export_excel: file_name={file_name}")
    
    mgr = _get_scraper_manager()
    all_jobs = getattr(mgr, '_last_search_results', [])
    
    if not all_jobs:
        return json.dumps({"success": False, "message": "没有可导出的岗位数据，请先搜索岗位。"}, ensure_ascii=False)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "岗位数据"
        
        # 表头
        headers = ["序号", "岗位名称", "公司名称", "薪资", "城市", "平台", "经验要求", "学历要求", "福利标签"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 数据行
        for idx, job in enumerate(all_jobs, 1):
            row_data = [
                idx,
                job.job_name or "",
                job.company_name or "",
                job.salary or "",
                job.location or "",
                job.platform or "",
                "",  # 经验要求（JobItem 无此字段）
                "",  # 学历要求（JobItem 无此字段）
                "",  # 福利标签（JobItem 无此字段）
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=value)
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
        
        # 调整列宽
        col_widths = [6, 30, 25, 15, 10, 12, 12, 12, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 保存文件
        if not file_name:
            file_name = f"岗位数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path = UPLOAD_DIR / f"{file_name}.xlsx"
        wb.save(str(file_path))
        
        return json.dumps({
            "success": True,
            "message": f"成功导出 {len(all_jobs)} 条岗位数据",
            "file_name": f"{file_name}.xlsx",
            "file_path": str(file_path),
            "count": len(all_jobs)
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] export_excel 失败: {e}")
        return json.dumps({"success": False, "message": f"导出失败: {str(e)}"}, ensure_ascii=False)


@tool
async def schedule_search(keyword: str, city: str, cron_expression: str, platforms: Optional[str] = None) -> str:
    """
    设置定时搜索任务。支持每天早上/每小时自动搜索岗位。
    
    参数:
        keyword: 搜索关键词
        city: 城市名称
        cron_expression: 定时表达式，如:
            - "每天8点" -> 每天早上8:00执行
            - "每小时" -> 每小时执行一次
            - "每周一9点" -> 每周一早上9:00执行
            - 或直接提供 cron 表达式，如 "0 8 * * *"（每天早上8点）
        platforms: 可选，指定平台
    """
    logger.info(f"[Agent Tool] schedule_search: keyword={keyword}, city={city}, cron={cron_expression}")
    
    try:
        from scheduler.scheduler import get_scheduler
        
        # 解析 cron 表达式
        cron_map = {
            "每小时": "0 * * * *",
            "每天8点": "0 8 * * *",
            "每天9点": "0 9 * * *",
            "每天10点": "0 10 * * *",
            "每天12点": "0 12 * * *",
            "每天18点": "0 18 * * *",
            "每周一9点": "0 9 * * 1",
            "每周一8点": "0 8 * * 1",
        }
        
        if cron_expression in cron_map:
            cron = cron_map[cron_expression]
        elif len(cron_expression.split()) == 5:
            cron = cron_expression  # 直接使用用户提供的 cron
        else:
            cron = "0 8 * * *"  # 默认每天早上8点
        
        scheduler = get_scheduler()
        job_id = scheduler.add_job(keyword, city, cron, platforms)
        
        # Guardrails: 定时任务审计
        get_guardrail_manager().log_schedule_operation("create", {
            "keyword": keyword, "city": city, "cron": cron, "platforms": platforms, "job_id": job_id
        })
        
        return json.dumps({
            "success": True,
            "message": f"定时搜索任务已创建。关键词: {keyword}, 城市: {city}, 定时: {cron_expression}",
            "job_id": job_id,
            "cron": cron,
            "keyword": keyword,
            "city": city,
            "platforms": platforms
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"[Agent Tool] schedule_search 失败: {e}")
        return json.dumps({"success": False, "message": f"创建定时任务失败: {str(e)}"}, ensure_ascii=False)


@tool
async def login_platform(platform: str) -> str:
    """
    打开招聘平台登录窗口（弹出Chrome浏览器），让用户手动扫码/输入账号登录。
    登录成功后自动保存Cookie，之后即可正常搜索该平台的岗位。
    
    参数:
        platform: 平台名称，可选: "boss"/"BOSS直聘", "liepin"/"猎聘", "51job"/"前程无忧", "zhaopin"/"智联招聘"
    
    注意：登录窗口会弹出Chrome浏览器，用户需在弹出窗口中完成登录操作（扫码或账号密码）。
    """
    logger.info("[Tools][login_platform] === 触发登录: {} ===", platform)
    
    # 平台名称映射
    PLATFORM_MAP = {
        "boss": "boss", "boss直聘": "boss",
        "liepin": "liepin", "猎聘": "liepin",
        "51job": "51job", "job51": "51job", "前程无忧": "51job",
        "zhaopin": "zhaopin", "智联": "zhaopin", "智联招聘": "zhaopin",
    }
    
    platform_key = PLATFORM_MAP.get(platform.lower(), platform)
    mgr = _get_scraper_manager()
    scraper = mgr.scrapers.get(platform_key)
    
    if not scraper:
        return json.dumps({
            "success": False,
            "message": f"未知平台: {platform}，支持: BOSS直聘、猎聘、前程无忧、智联招聘"
        }, ensure_ascii=False)
    
    if not hasattr(scraper, "manual_login"):
        return json.dumps({
            "success": False,
            "message": f"{getattr(scraper, 'platform', platform)} 不支持手动登录功能。"
        }, ensure_ascii=False)
    
    display_name = getattr(scraper, "platform", platform)
    
    try:
        loop = asyncio.get_running_loop()
        
        def _do_login():
            return asyncio.run(scraper.manual_login())
        
        logger.info("[Tools][login_platform] 正在打开 {} 登录窗口，请在弹出的Chrome中完成登录...", display_name)
        success, jobs = await loop.run_in_executor(None, _do_login)
        
        if success:
            logger.info("[Tools][login_platform] {} 登录成功!", display_name)
            # Guardrails: 登录审计
            get_guardrail_manager().log_login_attempt(display_name, True)
            return json.dumps({
                "success": True,
                "message": f"{display_name} 登录成功！Cookie已保存，现在可以正常搜索{display_name}的岗位了。"
            }, ensure_ascii=False)
        else:
            logger.warning("[Tools][login_platform] {} 登录未完成或超时", display_name)
            get_guardrail_manager().log_login_attempt(display_name, False)
            return json.dumps({
                "success": False,
                "message": f"{display_name} 登录未完成。请重试，或在弹出的浏览器窗口中检查是否遇到验证码。建议扫码登录更方便。"
            }, ensure_ascii=False)
    except Exception as e:
        logger.error("[Tools][login_platform] 登录异常: {}", e)
        return json.dumps({
            "success": False,
            "message": f"{display_name} 登录过程出错: {str(e)}"
        }, ensure_ascii=False)


@tool
async def auto_apply_jobs(config: RunnableConfig, max_count: int = 5, match_threshold: float = 0.5) -> str:
    """自动投递简历：对当前搜索到的岗位进行简历匹配，然后自动向HR发送打招呼消息。
    
    参数:
        max_count: 最多投递多少个匹配岗位，默认5个
        match_threshold: 最低匹配分数阈值（0-1），默认0.5
    
    使用前需确保：
    1. 已有搜索结果（先执行搜索）
    2. 已上传简历文件
    3. BOSS直聘已登录
    """
    logger.info("[Tools][auto_apply] === 开始自动投递 ===")
    logger.info("[Tools][auto_apply] 参数: max_count={}, match_threshold={}", max_count, match_threshold)
    
    mgr = _get_scraper_manager()
    jobs = getattr(mgr, '_last_search_results', [])
    
    if not jobs:
        return json.dumps({
            "success": False,
            "message": "没有搜索到岗位，请先使用搜索功能获取岗位列表。"
        }, ensure_ascii=False)
    
    logger.info("[Tools][auto_apply] 当前缓存岗位数: {}", len(jobs))
    
    # 查找简历文件（优先对话专属目录，确保对话隔离）
    thread_id = config.get("configurable", {}).get("thread_id", "default") if config else "default"
    resume_text = ""
    resume_path = ""
    resume_files = _find_resume_files(thread_id)
    if resume_files:
        resume_path = str(resume_files[0])
        try:
            resume_text = extract_resume_text(resume_path)
            logger.info("[Tools][auto_apply] 简历解析成功: {}, 文本长度={}", resume_files[0].name, len(resume_text))
        except Exception as e:
            logger.error("[Tools][auto_apply] 简历解析失败: {}", e)
            return json.dumps({
                "success": False,
                "message": f"简历解析失败: {e}。请确认已上传PDF/Word/TXT格式的简历文件。"
            }, ensure_ascii=False)
    
    if not resume_text:
        return json.dumps({
            "success": False,
            "message": "未找到简历文件。请先上传简历到上传目录。"
        }, ensure_ascii=False)
    
    # 简历匹配 - 使用关键词匹配评分
    logger.info("[Tools][auto_apply] 开始简历匹配...")
    
    # 从简历中提取技能关键词
    from agent.apply_manager import extract_keywords_from_resume
    resume_skills = extract_keywords_from_resume(resume_text)
    logger.info("[Tools][auto_apply] 简历技能关键词: {}", resume_skills)
    
    # 对每个岗位计算匹配分
    def calc_match_score(job: any, skills: list) -> float:
        """基于关键词匹配计算岗位匹配度评分"""
        if not skills:
            return 0.3  # 无技能信息时给一个基础分
        
        job_text = f"{job.job_name} {job.requirements} {job.responsibilities} {job.company_industry}".lower()
        
        matched = 0
        weight = 0
        for skill in skills:
            if skill.lower() in job_text:
                # 核心技能权重更高
                if skill.lower() in job.job_name.lower():
                    matched += 2
                    weight += 2
                else:
                    matched += 1
                    weight += 1
        
        if weight == 0:
            return 0.1
        
        # 归一化到 0-1
        score = matched / (len(skills) * 1.5)
        return min(score, 1.0)
    
    # 对所有有职位信息的岗位评分（按平台分组）
    platform_map = {
        "BOSS直聘": "boss",
        "猎聘": "liepin",
        "前程无忧": "51job",
        "智联招聘": "zhaopin",
    }
    
    # 收集各平台可投递的岗位（有 job_url 或 platform_job_id）
    all_eligible = []
    for j in jobs:
        if j.platform in platform_map and (j.job_url or j.platform_job_id):
            score = calc_match_score(j, resume_skills)
            all_eligible.append((score, j))
    
    logger.info("[Tools][auto_apply] 可投递岗位总数: {} (四平台)", len(all_eligible))
    
    if not all_eligible:
        return json.dumps({
            "success": False,
            "message": "当前搜索结果中没有可投递的岗位。请先执行搜索。"
        }, ensure_ascii=False)
    
    # 按匹配度排序
    all_eligible.sort(key=lambda x: x[0], reverse=True)
    
    # 过滤低于阈值的
    candidates = [(s, j) for s, j in all_eligible if s >= match_threshold]
    logger.info("[Tools][auto_apply] 匹配阈值以上岗位数: {}", len(candidates))
    
    # 限制数量
    candidates = candidates[:max_count]
    
    if not candidates:
        top_score = all_eligible[0][0] if all_eligible else 0
        return json.dumps({
            "success": True,
            "message": f"没有找到匹配度高于 {match_threshold} 的岗位。最高匹配分: {top_score:.2f}",
            "applied": 0,
            "total_suitable": len(all_eligible)
        }, ensure_ascii=False)
    
    # 打印候选岗位摘要
    for s, j in candidates:
        logger.info("[Tools][auto_apply] 候选: [{}] {} | {} | 匹配度={}", 
                   j.platform, j.job_name, j.company_name, round(s, 2))
    
    # Guardrails 第1层：投递前综合检查
    guard = get_guardrail_manager()
    ok, block_reason = guard.pre_apply_check(platform="", apply_count=len(candidates))
    if not ok:
        logger.warning("[Tools][auto_apply] Guardrails 拦截: {}", block_reason)
        return json.dumps({
            "success": False,
            "message": f"投递被安全策略拦截: {block_reason}",
            "guardrails_status": guard.get_rate_limit_status(),
        }, ensure_ascii=False)
    
    # 执行投递（四平台并行）
    logger.info("[Tools][auto_apply] 开始投递，候选岗位数: {}", len(candidates))
    results = []
    applied_count = 0
    skipped_count = 0
    
    async def _apply_one(score: float, job) -> dict:
        """对单个岗位执行投递"""
        nonlocal applied_count, skipped_count
        
        platform_id = job.platform_job_id or job.job_url
        if not platform_id:
            return {
                "platform": job.platform,
                "job_name": job.job_name,
                "company": job.company_name,
                "match_score": round(score, 2),
                "status": "failed",
                "reason": "缺少岗位标识"
            }
        
        # 检查是否已投递
        if is_already_applied(str(platform_id), job.platform):
            skipped_count += 1
            return {
                "platform": job.platform,
                "job_name": job.job_name,
                "company": job.company_name,
                "match_score": round(score, 2),
                "status": "skipped",
                "reason": "已投递过"
            }
        
        # 生成打招呼消息
        greeting = generate_greeting(job, resume_text, score)
        
        # Guardrails 第2层：招呼语内容审核
        guard = get_guardrail_manager()
        greeting_ok, greeting_reason = guard.validate_greeting(greeting)
        if not greeting_ok:
            return {
                "platform": job.platform,
                "job_name": job.job_name,
                "company": job.company_name,
                "match_score": round(score, 2),
                "status": "blocked",
                "reason": f"招呼语审核未通过: {greeting_reason}"
            }
        
        # 获取对应的爬虫
        scraper_key = platform_map.get(job.platform)
        if not scraper_key:
            return {
                "platform": job.platform,
                "job_name": job.job_name,
                "company": job.company_name,
                "match_score": round(score, 2),
                "status": "failed